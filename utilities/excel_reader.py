import openpyxl


def get_data_from_excel(path,sheetname):

    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheetname]

    values=[]

    for row in range(1,sheet.max_row+1):
        nested_values=[]
        for column in range(1,sheet.max_column+1):
            data=sheet.cell(row,column).value
            nested_values.append(data)
        values.append(nested_values)

    return values